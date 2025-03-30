import openpyxl
book = openpyxl.load_workbook("C:\\Users\\HP\\Desktop\\PythonDemo.xlsx")
sheet = book.active
cell=sheet.cell(row=1, column=2)
Dict={}
print(cell.value)

sheet.cell(row=1, column=2).value = "Prasad"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "Testcase2":

        for j in  range(2, sheet.max_column+1):
            Dict[sheet.cell(row=i, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)
